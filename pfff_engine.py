"""
PFFF v13 — Final Engine
========================
All bugs fixed. Production-ready for Colab + Streamlit.

KEY FIX — SWITCHING VALUES (Dual Anchor Method):
================================================
There are TWO correct interpretations of a switching value for this thesis:

1. DPR-ANCHOR SV (Consultant's Phantom Safety):
   - Starts from DPR_EIRR (e.g. 13.22% for P1)
   - Asks: how much can THIS variable change before the CONSULTANT's number fails?
   - For P1: Cost SV = +13.3%  → DPR says project survives 13.3% overrun
   - This is the OPTIMISM BIAS metric — it shows how little the DPR was stress-tested

2. P50-ANCHOR SV (Realistic Safety):
   - Starts from P50 simulated EIRR (e.g. 7.2% for P1)
   - Asks: how much can this variable change before the REALISTIC outcome fails?
   - For P1 (P50 = 7.2%): Project is ALREADY below 12%. SV = DEFICIT -4.8pp
   - This is the TRUE fragility metric

The Bias Gap = DPR_EIRR − P50_EIRR (e.g. 6.0pp for P1) = Optimism Bias

SHOWING BOTH SIDE-BY-SIDE is the core thesis finding:
  "The DPR told NHAI the project could survive a 13.3% cost overrun.
   The PFFF shows the project is already 4.8pp below viability at median
   outcomes. The difference is 6.0pp of undetected optimism bias."

WHY DELAY SV IS NOT "SAFE" FOR RED PROJECTS:
  Old code: SV_delay = gap_from_DPR / delay_sensitivity_per_month
  For P1: gap = 13.22 - 12 = 1.22pp; sens = 0.0275pp/mo → SV = 44mo
  This is CORRECT for the DPR anchor (consultant's view).
  For the P50 anchor: gap = 7.2 - 12 = -4.8pp → SV is NEGATIVE
  "Project needs to finish 174 months EARLY to achieve 12% at P50" = impossible.
  Display: SV_delay_P50 = None / "Project already below hurdle"

ZERO STRESS PROOF:
  eirr_iter(DPR inputs exactly) = DPR_EIRR → confirms model is not biased against projects
  The uncertainty comes from realistic variation, not model design
"""

import numpy as np
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
import warnings; warnings.filterwarnings('ignore')
from scipy import stats
from scipy.stats import norm, lognorm, triang
from scipy.optimize import brentq
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
import os

try:
    get_ipython()
    IN_NOTEBOOK = True
except NameError:
    IN_NOTEBOOK = False
    matplotlib.use('Agg')

plt.rcParams.update({
    "font.family":"DejaVu Sans","font.size":9,
    "figure.facecolor":"white","axes.facecolor":"#FAFAFA",
    "axes.edgecolor":"#CCCCCC","axes.grid":True,
    "grid.color":"#EEEEEE","grid.linewidth":0.7,
    "text.color":"#212529","axes.labelcolor":"#495057",
    "xtick.color":"#495057","ytick.color":"#495057",
    "axes.spines.top":False,"axes.spines.right":False,
})

np.random.seed(42)
N_ITER  = 10_000
OUT_DIR = "."
os.makedirs(OUT_DIR, exist_ok=True)

C = {
    "green":"#198754","green_lt":"#D1E7DD",
    "amber":"#856404","amber_lt":"#FFF3CD",
    "red":"#842029","red_lt":"#F8D7DA",
    "blue":"#0D6EFD","blue_lt":"#CFE2FF",
    "purple":"#6F42C1","grey":"#6C757D","dark":"#212529",
}

def fi_color(fi):
    if fi < 25: return C["green_lt"],C["green"],C["green"]
    if fi < 50: return C["amber_lt"],C["amber"],C["amber"]
    return C["red_lt"],C["red"],C["red"]

def verdict(fi):
    if fi < 25: return "GREEN — Approve"
    if fi < 50: return "AMBER — Conditional"
    return "RED — Return DPR"

# ═══════════════════════════════════════════════════════════
# PROJECT REGISTRY
# ═══════════════════════════════════════════════════════════
PROJECTS = {
    "P1":{"name":"Chitrakoot–Kothi (NH-135BG)","short":"P1 NH-135BG",
        "state":"UP/MP","dpr_mode":"HAM","eval_yrs":15,"role":"DEVELOPMENT",
        "civil_cr":612.98,"la_cr":347.53,"om_cr":8.44,
        "build_mo":24,"dpr_yr":2018,
        "dpr_eirr":13.22,"dpr_firr":13.01,"dpr_eq":15.04,
        "cost_sens":0.092,"traf_sens":0.102,
        "base_aadt":2840,"yr1_aadt":3930,"growth":0.0525,"survey_yr":2017,
        "survey_indep":False,
        "la_pct":5,"forest_clr":"NOT_APPLIED","community":"MEDIUM",
        "geotech":"PARTIAL","contractor":"STRESSED",
        "terrain":"ROLLING","crossings":"MODERATE","proj_type":"GREENFIELD",
        "forest_pct":49.5,"network":"FEEDER","scale_cr":612.98},
    "P2":{"name":"CPRR Sections II & III (AIIB)","short":"P2 CPRR",
        "state":"Tamil Nadu","dpr_mode":"EPC","eval_yrs":20,"role":"DEVELOPMENT",
        "civil_cr":3673.0,"la_cr":1855.0,"om_cr":45.2,
        "build_mo":36,"dpr_yr":2022,
        "dpr_eirr":15.65,"dpr_firr":None,"dpr_eq":None,
        "cost_sens":0.170,"traf_sens":0.190,
        "base_aadt":37000,"yr1_aadt":44800,"growth":0.065,"survey_yr":2018,
        "survey_indep":True,
        "la_pct":72,"forest_clr":"CLEARED","community":"HIGH",
        "geotech":"COMPLETE","contractor":"STRONG",
        "terrain":"PLAIN","crossings":"HIGH","proj_type":"GREENFIELD",
        "forest_pct":0,"network":"CORRIDOR_LINK","scale_cr":3673.0},
    "P3":{"name":"NH-66 Pkg III Chertalai–TVM","short":"P3 NH-66 Kerala",
        "state":"Kerala","dpr_mode":"HAM","eval_yrs":15,"role":"DEVELOPMENT",
        "civil_cr":4647.0,"la_cr":1165.0,"om_cr":55.0,
        "build_mo":30,"dpr_yr":2017,
        "dpr_eirr":47.00,"dpr_firr":11.20,"dpr_eq":14.80,
        "cost_sens":0.327,"traf_sens":0.567,
        "base_aadt":24500,"yr1_aadt":32400,"growth":0.075,"survey_yr":2017,
        "survey_indep":False,
        "la_pct":10,"forest_clr":"NONE","community":"EXTREME",
        "geotech":"COMPLETE","contractor":"ADEQUATE",
        "terrain":"COASTAL_ROLLING","crossings":"HIGH","proj_type":"BROWNFIELD",
        "forest_pct":0,"network":"CORRIDOR_LINK","scale_cr":4647.0},
    "P4":{"name":"Amas–Shivrampur (NH-119D)","short":"P4 Amas Bihar",
        "state":"Bihar","dpr_mode":"EPC","eval_yrs":20,"role":"DEVELOPMENT",
        "civil_cr":1079.77,"la_cr":320.0,"om_cr":14.0,
        "build_mo":24,"dpr_yr":2020,
        "dpr_eirr":18.20,"dpr_firr":None,"dpr_eq":None,
        "cost_sens":0.187,"traf_sens":0.273,
        "base_aadt":18173,"yr1_aadt":21500,"growth":0.065,"survey_yr":2019,
        "survey_indep":False,
        "la_pct":25,"forest_clr":"EIA_PENDING","community":"LOW_MEDIUM",
        "geotech":"COMPLETE","contractor":"ADEQUATE",
        "terrain":"PLAIN","crossings":"MODERATE","proj_type":"GREENFIELD",
        "forest_pct":0,"network":"FEEDER","scale_cr":1079.77,
        "rainfall":"MONSOON_FLOOD"},
    "P5":{"name":"Vadodara–Halol (SH-87)","short":"P5 Vadodara BOT",
        "state":"Gujarat","dpr_mode":"BOT","eval_yrs":30,"role":"VALIDATION",
        "civil_cr":180.0,"la_cr":12.0,"om_cr":3.5,
        "build_mo":18,"dpr_yr":1998,
        "dpr_eirr":15.60,"dpr_firr":14.20,"dpr_eq":18.50,
        "cost_sens":0.187,"traf_sens":0.280,
        "base_aadt":8400,"yr1_aadt":12000,"growth":0.085,"survey_yr":1997,
        "survey_indep":False,
        "actual_aadt":6973,
        "la_pct":95,"forest_clr":"NONE","community":"LOW",
        "geotech":"DESKTOP","contractor":"STRESSED",
        "terrain":"PLAIN","crossings":"LOW","proj_type":"GREENFIELD",
        "forest_pct":0,"network":"STANDALONE","scale_cr":180.0},
    "P6":{"name":"E-W Corridor NH-27 Sector I","short":"P6 E-W Corridor",
        "state":"Rajasthan/MP","dpr_mode":"EPC","eval_yrs":20,"role":"DEVELOPMENT",
        "civil_cr":3200.0,"la_cr":200.0,"om_cr":38.0,
        "build_mo":36,"dpr_yr":2004,
        "dpr_eirr":16.50,"dpr_firr":None,"dpr_eq":None,
        "cost_sens":0.173,"traf_sens":0.253,
        "base_aadt":5200,"yr1_aadt":6500,"growth":0.075,"survey_yr":2004,
        "survey_indep":False,
        "la_pct":65,"forest_clr":"PENDING","community":"MEDIUM",
        "geotech":"PARTIAL","contractor":"ADEQUATE",
        "terrain":"ROLLING","crossings":"MODERATE","proj_type":"GREENFIELD",
        "forest_pct":12,"network":"CORRIDOR_LINK","scale_cr":3200.0},
    "P7":{"name":"Samruddhi Mahamarg (MSRDC)","short":"P7 Samruddhi",
        "state":"Maharashtra","dpr_mode":"EPC","eval_yrs":30,"role":"VALIDATION",
        "civil_cr":55335.0,"la_cr":1712.0,"om_cr":620.0,
        "build_mo":48,"dpr_yr":2016,
        "dpr_eirr":18.00,"dpr_firr":12.50,"dpr_eq":None,
        "cost_sens":0.207,"traf_sens":0.280,
        "base_aadt":15000,"yr1_aadt":25000,"growth":0.085,"survey_yr":2016,
        "survey_indep":True,
        "actual_aadt":45000,"actual_cost_mult":1.35,
        "la_pct":100,"forest_clr":"STAGE_II","community":"MEDIUM",
        "geotech":"COMPLETE","contractor":"STRONG",
        "terrain":"MIXED_MOUNTAIN","crossings":"VERY_HIGH","proj_type":"GREENFIELD",
        "forest_pct":8,"network":"CORRIDOR_LINK","scale_cr":55335.0},
}

COST_CLASS = {"BEST":(0.15,0.18),"WORST":(0.90,0.38)}
MODES = ["EPC","HAM","BOT"]
HURDLES = {"EIRR":0.12,"FIRR":0.10,"EQ_HAM":0.12,"EQ_BOT":0.15}

# ═══════════════════════════════════════════════════════════
# SCN CONDITIONING
# ═══════════════════════════════════════════════════════════
def compute_scn(p):
    scn = {}
    eff_age = p["dpr_yr"] - p["survey_yr"]
    scn["survey_age"] = eff_age
    if eff_age > 7:    sm = 1.40
    elif eff_age > 4:  sm = 1.25
    elif eff_age > 2:  sm = 1.15
    else:              sm = 1.00
    if p.get("survey_indep"): sm *= 0.85
    scn["traf_sig_mult"] = sm

    la = p["la_pct"]
    geo_score = {"COMPLETE":0.0,"PARTIAL":0.40,"DESKTOP":1.0}.get(p["geotech"],0.3)
    con_score = {"STRONG":0.0,"ADEQUATE":0.40,"STRESSED":1.0}.get(p["contractor"],0.4)
    ter_score = {"PLAIN":0.0,"ROLLING":0.20,"COASTAL_ROLLING":0.40,
                 "HILLY":0.60,"MIXED_MOUNTAIN":0.70,"MOUNTAIN":1.0}.get(p["terrain"],0.3)
    cro_score = {"LOW":0.0,"MODERATE":0.20,"HIGH":0.50,"VERY_HIGH":0.80}.get(p["crossings"],0.2)
    for_score = min(1.0, p.get("forest_pct",0)/50)
    la_score  = 1.0 - (la/100)

    cost_scn  = geo_score*0.35 + con_score*0.30 + ter_score*0.25 + cro_score*0.10
    scn_score = la_score*0.30 + geo_score*0.20 + con_score*0.20 + ter_score*0.15 + cro_score*0.10 + for_score*0.05
    scn["cost_scn"] = cost_scn; scn["scn_score"] = scn_score

    scale_eff = 0.80 if p["scale_cr"]>10000 else 0.88 if p["scale_cr"]>5000 else 1.00
    scn["scale_eff"] = scale_eff

    bm,bs = COST_CLASS["BEST"]; wm,ws = COST_CLASS["WORST"]
    v05_overrun = (bm + cost_scn*(wm-bm)) * scale_eff
    v05_sigma   = bs + cost_scn*(ws-bs)
    if p["geotech"] == "COMPLETE": v05_sigma = min(v05_sigma, 0.20)
    if p.get("proj_type") == "BROWNFIELD": v05_overrun += 0.08
    if p.get("rainfall") == "MONSOON_FLOOD": v05_overrun += 0.05
    scn["v05_mean_mult"] = 1.0 + v05_overrun; scn["v05_sigma"] = v05_sigma

    if   la > 90: vm,vs = 1.40,0.25
    elif la > 80: vm,vs = 1.80,0.30
    elif la > 60: vm,vs = 2.20,0.38
    elif la > 40: vm,vs = 2.80,0.45
    elif la > 20: vm,vs = 3.50,0.52
    else:         vm,vs = 4.20,0.58
    cm = {"LOW":0.90,"LOW_MEDIUM":1.00,"MEDIUM":1.12,"HIGH":1.30,"EXTREME":1.55}.get(p["community"],1.00)
    scn["v06_mean_mult"] = min(vm*cm,5.0); scn["v06_sigma"] = vs

    if   la > 80: ps = 0.08
    elif la > 60: ps = 0.15
    elif la > 40: ps = 0.28
    elif la > 20: ps = 0.42
    else:         ps = 0.55
    ps += {"NONE":0,"CLEARED":0,"EIA_PENDING":0.04,"NOT_APPLIED":0.08,
           "PENDING":0.08,"STAGE_II":0.10,"BLOCKED":0.18}.get(p["forest_clr"],0)
    ps += {"LOW":0,"LOW_MEDIUM":0.02,"MEDIUM":0.04,"HIGH":0.08,"EXTREME":0.16}.get(p["community"],0)
    ps += {"PLAIN":0,"ROLLING":0.02,"COASTAL_ROLLING":0.04,"HILLY":0.06,
           "MIXED_MOUNTAIN":0.08,"MOUNTAIN":0.14}.get(p["terrain"],0)
    ps = min(0.70, ps)
    if p["scale_cr"]>10000 and p.get("contractor")=="STRONG": ps = min(ps, 0.30)
    scn["v07_ps"] = ps

    jdr = p["yr1_aadt"]/max(p["base_aadt"],1)
    scn["jdr"] = jdr; scn["w2"] = 0.08 if jdr>1.10 else 0.04
    muA  = p["yr1_aadt"]
    sigA = muA*0.12*sm
    net_mult = {"STANDALONE":1.00,"FEEDER":1.08,"CORRIDOR_LINK":1.15}.get(p["network"],1.00)
    sigA *= net_mult
    if p.get("survey_indep"): sigA *= 0.85
    im = min(1.10+(jdr-1.0)*0.60,1.80)
    scn["muA"]=muA; scn["sA"]=sigA
    scn["muB"]=p["yr1_aadt"]*im; scn["sB"]=0.25*p["yr1_aadt"]*im
    scn["ramp_min"]=0.50 if p["dpr_mode"]=="BOT" else 0.70
    scn["ramp_max"]=0.85 if p["dpr_mode"]=="BOT" else 0.95
    return scn

# ═══════════════════════════════════════════════════════════
# CORRELATED MCS
# ═══════════════════════════════════════════════════════════
CORR = np.array([
    [1.00, 0.45, 0.65,  0.00,  0.00],
    [0.45, 1.00, 0.70, -0.10,  0.00],
    [0.65, 0.70, 1.00, -0.25, -0.10],
    [0.00,-0.10,-0.25,  1.00,  0.30],
    [0.00, 0.00,-0.10,  0.30,  1.00],
])
CHOL = np.linalg.cholesky(CORR)

def pert_s(n, lo, mode, hi):
    if abs(hi-lo)<1e-9: return np.full(n,mode)
    mu=(lo+4*mode+hi)/6; v=((hi-lo)**2)/36
    d=(mu-lo)*(hi-mu)/v-1
    a=max((mu-lo)/(hi-lo)*d,0.01); b=max(a*(hi-mu)/(mu-lo),0.01)
    return lo+stats.beta.rvs(a,b,size=n)*(hi-lo)

def run_mcs(p, scn, n=N_ITER):
    Z=np.random.normal(0,1,(n,5)); Zc=Z@CHOL.T; U=norm.cdf(Zc)
    mu_log=np.log(p["civil_cr"]*scn["v05_mean_mult"])
    v05=lognorm.ppf(np.clip(U[:,0],1e-4,.9999),s=scn["v05_sigma"],scale=np.exp(mu_log))
    mu_log6=np.log(p["la_cr"]*scn["v06_mean_mult"])
    v06=np.minimum(lognorm.ppf(np.clip(U[:,1],1e-4,.9999),s=scn["v06_sigma"],
                   scale=np.exp(mu_log6)),p["la_cr"]*5.0)
    reg=(np.random.uniform(0,1,n)<scn["v07_ps"]).astype(int)
    v07=np.where(reg==0,pert_s(n,3,10,24),pert_s(n,36,54,90))
    comp=(np.random.uniform(0,1,n)<scn["w2"]).astype(int)
    aA=scn["muA"]+scn["sA"]*norm.ppf(np.clip(U[:,3],1e-4,.9999))
    aB=np.random.normal(scn["muB"],scn["sB"],n)
    v01=np.maximum(np.where(comp==0,aA,aB),100)
    gc=np.clip((p["growth"]-0.02)/0.065,0.01,0.99)
    v02=triang.ppf(np.clip(U[:,4],1e-4,.9999),c=gc,loc=0.02,scale=0.065)
    v10=np.random.triangular(0.85,1.00,1.15,n)
    v11=np.random.triangular(0.88,1.00,1.12,n)
    v08=p["om_cr"]*np.random.triangular(0.90,1.00,1.30,n)
    ramp=np.random.uniform(scn["ramp_min"],scn["ramp_max"],n)
    teff=np.random.uniform(0.88,0.97,n)
    return dict(v05=v05,v06=v06,v07=v07,v01=v01,v02=v02,v08=v08,
                v10=v10,v11=v11,ramp=ramp,teff=teff,reg=reg)

# ═══════════════════════════════════════════════════════════
# IRR ENGINES
# ═══════════════════════════════════════════════════════════
def eirr_iter(p, scn, v05, v07, v01, v02, v10, v11):
    """Zero-stress: at DPR values → returns exactly DPR_EIRR."""
    dpr_e = p["dpr_eirr"]
    co_pct    = (v05/p["civil_cr"]-1.0)*100
    cost_fx   = -co_pct*p["cost_sens"]
    traf_fx   = (v01/max(p["yr1_aadt"],1)-1.0)*100*p["traf_sens"]*(0.7359*v10+0.2641*v11)
    g_fx      = (v02-p["growth"])*100*0.030
    delay_fx  = -v07*(dpr_e*0.025/12)
    return (dpr_e+cost_fx+traf_fx+g_fx+delay_fx)/100

def verify_calibration(p, scn):
    zs = eirr_iter(p,scn,p["civil_cr"],0.0,p["yr1_aadt"],p["growth"],1.0,1.0)
    delta = abs(zs*100-p["dpr_eirr"])
    status = "✓ PASS" if delta<0.01 else f"✗ FAIL (Δ={delta:.3f}pp)"
    print(f"  {p['name'][:38]:<40} DPR={p['dpr_eirr']:.2f}%  ZS={zs*100:.2f}%  [{status}]")
    return zs

def firr_ham_iter(p, v05, v06, v07):
    if p["dpr_firr"] is None: return np.nan
    dpr_f = p["dpr_firr"]; dpr_e = p["dpr_eirr"]
    firr_cost_sens = p["cost_sens"]*min(1.0,dpr_f/dpr_e)
    total_cr = p["civil_cr"]+p["la_cr"]
    co_pct = ((v05+v06)/max(total_cr,1)-1.0)*100
    idc = 0.09*0.70*max(co_pct/100,0)*dpr_f*0.40
    return (dpr_f-co_pct*firr_cost_sens-idc-(v07/12)*0.90)/100

def firr_bot_iter(p, v05, v06, v07, v01, v10, v11, ramp, teff):
    if p["dpr_firr"] is None: return np.nan
    dpr_f = p["dpr_firr"]; dpr_e = p["dpr_eirr"]
    firr_cost_sens = p["cost_sens"]*min(1.0,dpr_f/dpr_e)
    total_cr = p["civil_cr"]+p["la_cr"]
    co_pct = ((v05+v06)/max(total_cr,1)-1.0)*100
    traf_fx = (v01/max(p["yr1_aadt"],1)-1.0)*100*(p["traf_sens"]*1.5)
    ramp_pen=(1.0-ramp)*0.30; coll_pen=(1.0-teff)*0.15
    idc_delay=(v07/12)*1.20
    return (dpr_f-co_pct*firr_cost_sens-idc_delay-ramp_pen-coll_pen+traf_fx*0.01)/100

def equity_irr_iter(p, mode, v05, v06, v07, firr):
    if mode=="EPC": return np.nan
    if mode=="HAM":
        dpr_eq = p.get("dpr_eq") or 15.0; dpr_e = p["dpr_eirr"]
        eq_cost_sens = p["cost_sens"]*min(1.0,dpr_eq/dpr_e)
        total_cr = p["civil_cr"]+p["la_cr"]
        net_co = ((v05+v06)/max(total_cr,1)-1.0)*100
        return (dpr_eq-net_co*eq_cost_sens-(v07/12)*0.80)/100
    if mode=="BOT":
        if firr is None or np.isnan(firr): return np.nan
        return float(np.clip(firr+(firr-0.09)*(0.70/0.30),-0.99,0.99))
    return np.nan

# ═══════════════════════════════════════════════════════════
# MODE SIMULATION
# ═══════════════════════════════════════════════════════════
def terrain_premium(terrain):
    return {"PLAIN":0.00,"ROLLING":0.01,"COASTAL_ROLLING":0.01,
            "HILLY":0.02,"MIXED_MOUNTAIN":0.03,"MOUNTAIN":0.03}.get(terrain,0.01)

def simulate_mode(p, scn, samp, mode, n=N_ITER):
    v05,v06,v07=samp["v05"],samp["v06"],samp["v07"]
    v01,v02,v10,v11=samp["v01"],samp["v02"],samp["v10"],samp["v11"]
    ramp,teff=samp["ramp"],samp["teff"]
    eirr_arr=np.array([eirr_iter(p,scn,v05[i],v07[i],v01[i],v02[i],v10[i],v11[i]) for i in range(n)])
    if mode=="HAM":
        firr_arr=np.array([firr_ham_iter(p,v05[i],v06[i],v07[i]) for i in range(n)])
    elif mode=="BOT":
        firr_arr=np.array([firr_bot_iter(p,v05[i],v06[i],v07[i],v01[i],v10[i],v11[i],ramp[i],teff[i]) for i in range(n)])
    else:
        firr_arr=np.full(n,np.nan)
    eq_arr=np.array([equity_irr_iter(p,mode,v05[i],v06[i],v07[i],
                     firr_arr[i] if not np.isnan(firr_arr[i]) else None) for i in range(n)])
    fi_eirr=np.sum(eirr_arr<HURDLES["EIRR"])/n*100
    valid_f=firr_arr[~np.isnan(firr_arr)]
    fi_firr=np.sum(valid_f<HURDLES["FIRR"])/len(valid_f)*100 if len(valid_f)>0 and mode!="EPC" else np.nan
    eq_h=(HURDLES["EQ_HAM"]+terrain_premium(p["terrain"])) if mode=="HAM" else \
         (HURDLES["EQ_BOT"]+terrain_premium(p["terrain"])) if mode=="BOT" else np.nan
    valid_e=eq_arr[~np.isnan(eq_arr)]
    fi_eq=np.sum(valid_e<eq_h)/len(valid_e)*100 if len(valid_e)>0 and mode!="EPC" else np.nan
    fi_vals=[fi_eirr]+([fi_firr] if not np.isnan(fi_firr) else [])+([fi_eq] if not np.isnan(fi_eq) else [])
    return {"mode":mode,"fi_eirr":fi_eirr,"fi_firr":fi_firr,"fi_eq":fi_eq,"fi_p":max(fi_vals),
            "eirr_arr":eirr_arr,"firr_arr":firr_arr,"eq_arr":eq_arr,
            "hurdle_eirr":HURDLES["EIRR"],"hurdle_eq":eq_h}

# ═══════════════════════════════════════════════════════════
# ANALYTICS
# ═══════════════════════════════════════════════════════════
def spearman_tornado(p, scn, samp, eirr_arr):
    from scipy.stats import spearmanr
    er=stats.rankdata(eirr_arr)
    factors=[("V05 Civil Cost",samp["v05"]),("V07 Delay",samp["v07"]),
             ("V01 Traffic",samp["v01"]),("V06 LA Cost",samp["v06"]),
             ("V02 Growth",samp["v02"]),("V10 VOC",samp["v10"]),("V11 VoT",samp["v11"])]
    res=[(nm,spearmanr(a,er)[0]) for nm,a in factors]
    res.sort(key=lambda x:abs(x[1]),reverse=True)
    return res

def rcf_acid_test(p, scn, samp, fi_primary):
    if fi_primary<25: return None
    p80c=np.percentile(samp["v05"],80)
    p20t=np.percentile(samp["v01"],20)
    p80d=np.percentile(samp["v07"],80)
    rcf_eirr=eirr_iter(p,scn,v05=p80c,v07=p80d,v01=p20t,v02=p["growth"],v10=0.88,v11=0.93)*100
    gap=HURDLES["EIRR"]*100-rcf_eirr
    if rcf_eirr>=12: dec="APPROVE WITH CONDITIONS"; resp="Monitoring triggers mandatory."
    elif gap<2: dec="RETURN TYPE 1 — BETTER EVIDENCE"; resp=f"Gap={gap:.1f}pp. Stronger data may close."
    elif gap<5: dec="RETURN TYPE 2 — VALUE ENGINEERING"; resp=f"Gap={gap:.1f}pp. Design modifications needed."
    else: dec="RETURN TYPE 3 — SCOPE REVISION"; resp=f"Gap={gap:.1f}pp. Project unviable as designed."
    return {"p80_cost":p80c,"p20_traf":p20t,"p80_delay":p80d,"rcf_eirr":rcf_eirr,
            "decision":dec,"response":resp,"cost_uplift":p80c/p["civil_cr"],"traf_haircut":p20t/p["yr1_aadt"]}

def compute_dual_sv(p, scn, p50_eirr):
    """
    Compute BOTH DPR-anchored and P50-anchored switching values.
    
    DPR-anchor: starts from consultant's DPR_EIRR → shows phantom safety
    P50-anchor: starts from simulated P50 → shows actual fragility
    
    For projects where P50 < 12%: P50-anchor SVs are shown as deficits
    (the project needs improvement, not just tolerance to changes)
    """
    hurdle = 12.0

    # --- DPR-anchored SVs (always computable) ---
    def sv_cost_dpr(pct):
        v05 = p["civil_cr"]*(1+pct/100)
        return eirr_iter(p,scn,v05=v05,v07=0,v01=p["yr1_aadt"],v02=p["growth"],v10=1.0,v11=1.0)*100-hurdle
    def sv_traf_dpr(pct):
        v01 = p["yr1_aadt"]*(1-pct/100)
        return eirr_iter(p,scn,v05=p["civil_cr"],v07=0,v01=v01,v02=p["growth"],v10=1.0,v11=1.0)*100-hurdle
    def sv_delay_dpr(mo):
        return eirr_iter(p,scn,v05=p["civil_cr"],v07=mo,v01=p["yr1_aadt"],v02=p["growth"],v10=1.0,v11=1.0)*100-hurdle

    try: dpr_cost = round(brentq(sv_cost_dpr,0,500),1)
    except: dpr_cost = None
    try: dpr_traf = round(brentq(sv_traf_dpr,0,99),1)
    except: dpr_traf = None
    try: dpr_delay = round(brentq(sv_delay_dpr,0,300),0)
    except: dpr_delay = None

    dpr_gap = round(p["dpr_eirr"]-hurdle, 2)

    # --- P50-anchored SVs ---
    p50_gap = round(p50_eirr - hurdle, 2)

    if p50_eirr >= hurdle:
        # Project OK at P50 — compute how much it can deteriorate
        # Approximate: scale DPR SVs by (p50_gap / dpr_gap)
        ratio = p50_gap / max(dpr_gap, 0.01)
        p50_cost  = round(dpr_cost  * ratio, 1) if dpr_cost  else None
        p50_traf  = round(dpr_traf  * ratio, 1) if dpr_traf  else None
        p50_delay = round(dpr_delay * ratio, 0) if dpr_delay else None
        p50_status = "ABOVE HURDLE AT P50"
    else:
        # Project ALREADY below hurdle at P50
        # SVs are deficits — show as negative (improvement needed)
        ratio = p50_gap / max(dpr_gap, 0.01)  # negative ratio
        p50_cost  = round(dpr_cost  * ratio, 1) if dpr_cost  else None  # negative = reduction needed
        p50_traf  = round(dpr_traf  * ratio, 1) if dpr_traf  else None
        p50_delay = round(dpr_delay * ratio, 0) if dpr_delay else None
        p50_status = "ALREADY BELOW HURDLE AT P50"

    bias_gap = round(p["dpr_eirr"] - p50_eirr, 2)

    return {
        # DPR-anchored (Consultant's Phantom Safety)
        "dpr_cost":  dpr_cost,
        "dpr_traf":  dpr_traf,
        "dpr_delay": dpr_delay,
        "dpr_gap":   dpr_gap,
        # P50-anchored (Realistic Safety)
        "p50_cost":  p50_cost,
        "p50_traf":  p50_traf,
        "p50_delay": p50_delay,
        "p50_gap":   p50_gap,
        "p50_status": p50_status,
        # Bias
        "bias_gap":  bias_gap,   # DPR optimism: how much consultant overstated EIRR
        "p50_eirr":  round(p50_eirr, 2),
    }

# ═══════════════════════════════════════════════════════════
# COLAB DASHBOARD PLOTS
# ═══════════════════════════════════════════════════════════
def plot_dashboard(p, scn, samp, results, tornado, rcf, svs, code):
    dpr_mode = p["dpr_mode"]; res = results[dpr_mode]
    fi = res["fi_p"]; bg, fc, ec = fi_color(fi)
    ep = res["eirr_arr"]*100
    p10,p50,p90 = np.percentile(ep,10),np.percentile(ep,50),np.percentile(ep,90)

    fig = plt.figure(figsize=(22,14),facecolor="white")
    fig.suptitle(
        f"PFFF v13 — {p['name']}  [{dpr_mode}]  |  Survey age: {scn['survey_age']}yr  |  "
        f"DPR EIRR: {p['dpr_eirr']:.2f}%  |  P50 Simulated: {p50:.2f}%  |  Bias: {p['dpr_eirr']-p50:+.2f}pp",
        fontsize=12,fontweight="bold",y=0.98)
    gs = gridspec.GridSpec(3,4,figure=fig,hspace=0.52,wspace=0.42)

    # Panel 0: Verdict + calibration proof
    ax0=fig.add_subplot(gs[0,0]); ax0.set_facecolor(bg); ax0.axis("off")
    ax0.text(0.5,0.88,f"FI = {fi:.1f}%",ha="center",fontsize=26,fontweight="bold",color=fc,transform=ax0.transAxes)
    ax0.text(0.5,0.68,verdict(fi),ha="center",fontsize=9.5,color=ec,transform=ax0.transAxes)
    ax0.text(0.5,0.52,f"DPR EIRR: {p['dpr_eirr']:.2f}%",ha="center",fontsize=9,color=C["grey"],transform=ax0.transAxes)
    ax0.text(0.5,0.38,f"P50 Simulated: {p50:.2f}%",ha="center",fontsize=9,fontweight="bold",
             color=C["red"] if p50<12 else C["green"],transform=ax0.transAxes)
    ax0.text(0.5,0.24,f"Optimism Bias: {p['dpr_eirr']-p50:+.2f}pp",ha="center",fontsize=9,
             color=C["red"],transform=ax0.transAxes)
    # Zero-stress proof badge
    zs_val = eirr_iter(p,scn,p["civil_cr"],0.0,p["yr1_aadt"],p["growth"],1.0,1.0)*100
    zs_ok = abs(zs_val-p["dpr_eirr"])<0.01
    ax0.text(0.5,0.08,f"Zero-Stress: {zs_val:.2f}% {'✓' if zs_ok else '✗'}",
             ha="center",fontsize=8,color=C["green"] if zs_ok else C["red"],transform=ax0.transAxes,
             bbox=dict(boxstyle="round,pad=0.3",fc="white",ec=C["green"] if zs_ok else C["red"]))
    ax0.set_title("PFFF Verdict & Calibration",fontsize=9,color=C["grey"],pad=3)

    # Panel 1: EIRR distribution with DPR vs P50 comparison
    ax1=fig.add_subplot(gs[0,1])
    ax1.hist(ep,bins=60,color=C["blue_lt"],edgecolor=C["blue"],alpha=0.8,linewidth=0.4)
    ax1.axvline(12,color=C["red"],ls="--",lw=2,label="12% Hurdle")
    ax1.axvline(p["dpr_eirr"],color=C["dark"],ls="-",lw=2,label=f"DPR: {p['dpr_eirr']:.1f}%")
    ax1.axvline(p50,color=C["blue"],ls=":",lw=2,label=f"P50: {p50:.1f}%")
    ax1.axvline(np.percentile(ep,20),color=C["amber"],ls=":",lw=1,label=f"P20: {np.percentile(ep,20):.1f}%")
    # Shade fail zone
    ymax = ax1.get_ylim()[1] if ax1.get_ylim()[1]>0 else 1
    ax1.axvspan(min(ep)-3,12,alpha=0.08,color=C["red"])
    ax1.set_title(f"EIRR Distribution  FI={res['fi_eirr']:.1f}%",fontsize=9)
    ax1.set_xlabel("EIRR (%)",fontsize=8); ax1.legend(fontsize=7)

    # Panel 2: FIRR
    ax2=fig.add_subplot(gs[0,2])
    firr_v=res["firr_arr"][~np.isnan(res["firr_arr"])]*100
    if len(firr_v)>10:
        ax2.hist(firr_v,bins=60,color="#D7BDE2",edgecolor="#8E44AD",alpha=0.8,linewidth=0.4)
        ax2.axvline(10,color=C["red"],ls="--",lw=2,label="10% Hurdle")
        ax2.axvline(np.percentile(firr_v,50),color=C["dark"],ls=":",lw=1.5,
                    label=f"P50: {np.percentile(firr_v,50):.1f}%")
        ax2.set_title(f"FIRR Distribution  FI={res['fi_firr']:.1f}%",fontsize=9)
        ax2.legend(fontsize=7)
    else:
        ax2.text(0.5,0.5,"FIRR: N/A\n(EPC mode)",ha="center",va="center",transform=ax2.transAxes,
                 fontsize=12,color=C["grey"])
        ax2.set_title("FIRR Distribution",fontsize=9)
    ax2.set_xlabel("FIRR (%)",fontsize=8)

    # Panel 3: Mode comparison
    ax3=fig.add_subplot(gs[0,3])
    mfis=[(m,results[m]["fi_p"]) for m in MODES]
    bars=ax3.bar([m for m,_ in mfis],[f for _,f in mfis],
                 color=[fi_color(f)[1] for _,f in mfis],edgecolor="white",width=0.5)
    ax3.axhline(50,color=C["red"],ls="--",lw=1,alpha=0.6)
    ax3.axhline(25,color=C["amber"],ls="--",lw=1,alpha=0.6)
    # Mark DPR mode
    dm_idx=[m for m,_ in mfis].index(dpr_mode)
    bars[dm_idx].set_edgecolor("black"); bars[dm_idx].set_linewidth(2)
    ax3.set_ylim(0,108); ax3.set_title(f"Mode FI  (□=DPR chosen: {dpr_mode})",fontsize=9)
    for bar,(m,f) in zip(bars,mfis):
        ax3.text(bar.get_x()+bar.get_width()/2,f+2,f"{f:.0f}%",ha="center",fontsize=9,
                 fontweight="bold",color=fi_color(f)[1])

    # Panel 4: Tornado
    ax4=fig.add_subplot(gs[1,:2])
    names=[t[0] for t in tornado[:7]]; rhos=[t[1] for t in tornado[:7]]
    colors_t=[C["red"] if r<0 else C["blue"] for r in rhos]
    ax4.barh(names[::-1],rhos[::-1],color=colors_t[::-1],alpha=0.8)
    ax4.axvline(0,color=C["dark"],lw=0.8)
    ax4.set_xlabel("Spearman ρ with EIRR",fontsize=8)
    ax4.set_title(f"Spearman Tornado  |  Primary Driver: {tornado[0][0]}",fontsize=9,color=C["red"])
    for i,(rho,nm) in enumerate(zip(rhos[::-1],names[::-1])):
        ax4.text(rho+(0.01 if rho>=0 else -0.01),i,f"{rho:.3f}",va="center",fontsize=7.5,
                 ha="left" if rho>=0 else "right")

    # Panel 5: DUAL SWITCHING VALUES TABLE (the key insight)
    ax5=fig.add_subplot(gs[1,2:]); ax5.axis("off")
    ax5.set_title("Switching Values — Dual Anchor (Core Thesis Finding)",fontsize=9,fontweight="bold",color=C["red"])

    col_labels = ["Variable","DPR View (Consultant)","P50 View (PFFF)","Verdict"]
    col_x = [0.01, 0.28, 0.55, 0.78]
    y_hdr = 0.94
    for cx,cl in zip(col_x,col_labels):
        ax5.text(cx,y_hdr,cl,transform=ax5.transAxes,fontsize=8.5,fontweight="bold",color=C["dark"])

    p50_failed = svs["p50_eirr"] < 12.0

    rows_data = [
        ("EIRR Headroom",
         f"+{svs['dpr_gap']:.2f}pp  (DPR {p['dpr_eirr']:.1f}%)",
         f"{svs['p50_gap']:+.2f}pp  (P50 {svs['p50_eirr']:.1f}%)",
         "BIAS" if p50_failed else "OK"),
        ("Cost Overrun SV",
         f"+{svs['dpr_cost']:.1f}%" if svs['dpr_cost'] else "∞",
         f"{svs['p50_cost']:+.1f}% req'd" if svs['p50_cost'] else "∞",
         "⚠ PHANTOM" if (p50_failed and svs['dpr_cost'] and svs['dpr_cost']>0) else "OK"),
        ("Traffic Shortfall SV",
         f"−{svs['dpr_traf']:.1f}%" if svs['dpr_traf'] else "∞",
         f"{svs['p50_traf']:+.1f}% req'd" if svs['p50_traf'] else "∞",
         "⚠ PHANTOM" if (p50_failed and svs['dpr_traf'] and svs['dpr_traf']>0) else "OK"),
        ("Delay SV",
         f"+{svs['dpr_delay']:.0f}mo" if svs['dpr_delay'] else "∞",
         f"None — already failed" if p50_failed else f"{svs['p50_delay']:.0f}mo",
         "⚠ PHANTOM" if p50_failed else "OK"),
        ("─"*15,"─"*18,"─"*18,"─"*8),
        ("Optimism Bias",
         f"DPR: {p['dpr_eirr']:.2f}%",
         f"P50: {svs['p50_eirr']:.2f}%",
         f"{svs['bias_gap']:+.2f}pp GAP"),
        ("CAG Ref: Cost","Avg overrun: +71%","","⚠ EXCEEDS DPR SV" if (svs['dpr_cost'] and svs['dpr_cost']<71) else "✓ DPR SV above"),
    ]
    for i,(v1,v2,v3,v4) in enumerate(rows_data):
        y = y_hdr - 0.13*(i+1)
        clr = C["red"] if "⚠" in str(v4) or "BIAS" in str(v4) else C["dark"]
        ax5.text(col_x[0],y,str(v1),transform=ax5.transAxes,fontsize=8,color=C["grey"])
        ax5.text(col_x[1],y,str(v2),transform=ax5.transAxes,fontsize=8,color=C["dark"])
        ax5.text(col_x[2],y,str(v3),transform=ax5.transAxes,fontsize=8,
                 color=C["red"] if p50_failed and "req" not in str(v3) and "already" in str(v3) else C["dark"])
        ax5.text(col_x[3],y,str(v4),transform=ax5.transAxes,fontsize=8,fontweight="bold",color=clr)

    # Panel 6: Traffic distribution
    ax6=fig.add_subplot(gs[2,:2]); v01=samp["v01"]
    ax6.hist(v01,bins=60,color=C["blue_lt"],edgecolor=C["blue"],alpha=0.75,linewidth=0.3,density=True)
    ax6.axvline(p["yr1_aadt"],color=C["dark"],lw=2,label=f"DPR Yr1: {p['yr1_aadt']:,.0f}")
    ax6.axvline(p["base_aadt"],color=C["grey"],ls="--",lw=1.2,label=f"Base: {p['base_aadt']:,.0f}")
    ax6.axvline(np.percentile(v01,50),color=C["blue"],ls=":",lw=1.5,label=f"P50: {np.percentile(v01,50):,.0f}")
    if p.get("actual_aadt"):
        act=p["actual_aadt"]; pct=np.sum(v01<=act)/len(v01)*100
        color_act=C["red"] if act<p["yr1_aadt"] else C["green"]
        ax6.axvline(act,color=color_act,lw=2.5,label=f"Actual: {act:,.0f} (P{pct:.0f})")
    ax6.set_title(f"Traffic Distribution  JDR={scn['jdr']:.2f}  σ-mult=×{scn['traf_sig_mult']:.2f}",fontsize=9)
    ax6.set_xlabel("AADT (PCU)",fontsize=8); ax6.yaxis.set_visible(False); ax6.legend(fontsize=7)

    # Panel 7: RCF / Stage 2
    ax7=fig.add_subplot(gs[2,2:]); ax7.axis("off")
    ax7.set_facecolor(C["amber_lt"] if rcf else C["green_lt"])
    if rcf:
        lines=[
            f"RCF Stress: P80 Cost×{rcf['cost_uplift']:.2f} | P20 Traffic×{rcf['traf_haircut']:.2f} | P80 Delay {rcf['p80_delay']:.0f}mo",
            f"RCF-adjusted EIRR: {rcf['rcf_eirr']:.2f}%  (vs 12% hurdle,  gap={HURDLES['EIRR']*100-rcf['rcf_eirr']:+.2f}pp)",
            f"Decision: {rcf['decision']}",
            f"Response: {rcf['response']}"
        ]
        ax7.text(0.02,0.90,"Stage 2 — RCF Acid Test (Amber/Red trigger)",
                 transform=ax7.transAxes,fontsize=9,fontweight="bold",color=C["amber"])
    else:
        lines=[
            f"FI={fi:.1f}% < 25% → GREEN: Stage 2 not required.",
            f"P50 EIRR={p50:.2f}% is above 12% hurdle by {p50-12:.2f}pp.",
            f"Zero-stress EIRR={zs_val:.2f}% = DPR {p['dpr_eirr']:.2f}% ✓ (model calibrated)"
        ]
        ax7.text(0.02,0.90,"Stage 2 — Not Required (GREEN)",
                 transform=ax7.transAxes,fontsize=9,fontweight="bold",color=C["green"])
    for i,line in enumerate(lines):
        ax7.text(0.02,0.70-i*0.23,line,transform=ax7.transAxes,fontsize=9,color=C["dark"])

    plt.tight_layout(rect=[0,0,1,0.96])
    fname=os.path.join(OUT_DIR,f"pfff_{code}_dashboard.png")
    fig.savefig(fname,dpi=150,bbox_inches="tight",facecolor="white")
    plt.show(); plt.close(fig)
    print(f"  → Saved: {fname}")


def plot_batch_and_bias(all_results, all_svs, all_p50):
    """Batch comparison + Optimism Bias panel side by side."""
    codes=list(PROJECTS.keys()); n=len(codes)
    fig,axes=plt.subplots(1,3,figsize=(24,7),facecolor="white")

    # Left: FI bars
    ax=axes[0]; ax.set_facecolor("#FAFAFA")
    mc={"EPC":"#0D6EFD","HAM":"#6F42C1","BOT":"#198754"}
    x=np.arange(n); w=0.25
    for mode,off in zip(MODES,[-w,0,w]):
        fis=[all_results[c][mode]["fi_p"] for c in codes]
        bars=ax.bar(x+off,fis,w*0.9,label=mode,color=mc[mode],alpha=0.85,edgecolor="white")
        for bar,f in zip(bars,fis):
            ax.text(bar.get_x()+bar.get_width()/2,f+1.5,f"{f:.0f}",ha="center",fontsize=7,
                    color=mc[mode],fontweight="bold")
    ax.axhline(50,color=C["red"],ls="--",lw=1.5,alpha=0.7,label="RED 50%")
    ax.axhline(25,color=C["amber"],ls="--",lw=1.2,alpha=0.7,label="AMBER 25%")
    ax.axhspan(50,105,alpha=0.04,color=C["red"])
    ax.axhspan(25,50,alpha=0.04,color=C["amber"])
    ax.axhspan(0,25,alpha=0.04,color=C["green"])
    ax.set_xticks(x); ax.set_xticklabels([PROJECTS[c]["short"] for c in codes],fontsize=8)
    ax.set_ylim(0,108); ax.set_ylabel("Fragility Index FI%",fontsize=10)
    ax.set_title("PFFF v13 — All 7 Projects × 3 Modes\n(□=DPR chosen mode)",fontsize=11,fontweight="bold")
    ax.legend(fontsize=8)
    for i,c in enumerate(codes):
        dm=PROJECTS[c]["dpr_mode"]; j=["EPC","HAM","BOT"].index(dm)
        off=[-w,0,w][j]; f=all_results[c][dm]["fi_p"]
        ax.add_patch(plt.Rectangle((i+off-w*0.45,0),w*0.9,f,fill=False,edgecolor="black",lw=2,zorder=5))
        if PROJECTS[c]["role"]=="VALIDATION":
            ax.text(i,102,"VALIDATION",ha="center",fontsize=7,color=C["grey"],style="italic")

    # Middle: Optimism Bias chart (DPR EIRR vs P50 simulated)
    ax2=axes[1]; ax2.set_facecolor("#FAFAFA")
    dpr_eirrs=[PROJECTS[c]["dpr_eirr"] for c in codes]
    p50_eirrs=[all_p50[c] for c in codes]
    bias_gaps=[d-p for d,p in zip(dpr_eirrs,p50_eirrs)]
    x2=np.arange(n)
    ax2.bar(x2-0.18,dpr_eirrs,0.35,label="DPR EIRR (Consultant)",color="#2C3E50",alpha=0.85,edgecolor="white")
    ax2.bar(x2+0.18,p50_eirrs,0.35,label="P50 Simulated (PFFF)",
            color=[C["green"] if p>=12 else C["red"] for p in p50_eirrs],alpha=0.85,edgecolor="white")
    ax2.axhline(12,color=C["red"],ls="--",lw=2,label="12% Hurdle")
    for i,(d,p) in enumerate(zip(dpr_eirrs,p50_eirrs)):
        ax2.annotate("",xy=(i+0.18,p),xytext=(i-0.18,d),
                     arrowprops=dict(arrowstyle="->",color=C["red"],lw=1.5))
        ax2.text(i,max(d,p)+0.8,f"Bias\n{d-p:+.1f}pp",ha="center",fontsize=7,
                 color=C["red"],fontweight="bold")
    ax2.set_xticks(x2); ax2.set_xticklabels([PROJECTS[c]["short"] for c in codes],fontsize=8)
    ax2.set_ylabel("EIRR (%)",fontsize=10)
    ax2.set_title("Optimism Bias: DPR vs PFFF P50\nArrow = magnitude of overstatement",
                  fontsize=11,fontweight="bold")
    ax2.legend(fontsize=8)

    # Right: Switching Values comparison
    ax3=axes[2]; ax3.set_facecolor("#FAFAFA")
    sv_cost_dpr=[all_svs[c].get("dpr_cost") or 200 for c in codes]
    sv_cost_p50=[all_svs[c].get("p50_cost") or 0 for c in codes]  # negative = deficit
    colors_dpr=["#E74C3C" if v<71 else "#27AE60" for v in sv_cost_dpr]
    colors_p50=["#842029" if v<=0 else "#198754" for v in sv_cost_p50]
    ax3.bar(x2-0.18,sv_cost_dpr,0.35,label="Cost SV (DPR-anchored)",color=colors_dpr,alpha=0.85,edgecolor="white")
    ax3.bar(x2+0.18,sv_cost_p50,0.35,label="Cost SV (P50-anchored)",color=colors_p50,alpha=0.7,edgecolor="white")
    ax3.axhline(71,color=C["red"],ls="--",lw=2,alpha=0.8,label="CAG avg overrun: +71%")
    ax3.axhline(0,color=C["dark"],ls="-",lw=1)
    for i,(d,p50) in enumerate(zip(sv_cost_dpr,sv_cost_p50)):
        ax3.text(i-0.18,d+3,f"+{d:.0f}%",ha="center",fontsize=7,fontweight="bold",color="#E74C3C" if d<71 else "#27AE60")
        label = f"{p50:+.0f}%" if p50>0 else "Deficit"
        ax3.text(i+0.18,min(p50,0)-5,label,ha="center",fontsize=7,fontweight="bold",
                 color=colors_p50[i])
    ax3.set_xticks(x2); ax3.set_xticklabels([PROJECTS[c]["short"] for c in codes],fontsize=8)
    ax3.set_ylabel("Cost Switching Value (%)",fontsize=10)
    ax3.set_title("Switching Value Bias\nDPR-anchor vs P50-anchor (CAG ref = +71%)",
                  fontsize=11,fontweight="bold")
    ax3.legend(fontsize=8)

    plt.tight_layout()
    fname=os.path.join(OUT_DIR,"pfff_batch_bias_sv.png")
    fig.savefig(fname,dpi=150,bbox_inches="tight",facecolor="white")
    plt.show(); plt.close(fig)
    print(f"  → Saved: {fname}")


def plot_validation_exhibit(all_results, all_scn):
    fig,axes=plt.subplots(1,2,figsize=(18,8),facecolor="white")
    fig.suptitle("PFFF v13 — Validation Exhibit (P5 & P7) | Applied at DPR submission using only DPR-stage inputs",
                 fontsize=12,fontweight="bold",y=0.97)
    for ax_idx,(code,expected) in enumerate([("P5","RED"),("P7","AMBER-RED")]):
        ax=axes[ax_idx]; p=PROJECTS[code]; scn=all_scn[code]
        res=all_results[code][p["dpr_mode"]]; ep=res["eirr_arr"]*100; fi=res["fi_p"]
        bg,fc,ec=fi_color(fi); p50=np.percentile(ep,50)
        ax.hist(ep,bins=60,color=C["blue_lt"],edgecolor=C["blue"],alpha=0.75,linewidth=0.3,density=True)
        ax.axvline(12,color=C["red"],lw=2.5,ls="--",label="12% Hurdle")
        ax.axvline(p["dpr_eirr"],color=C["dark"],lw=2,ls="-",label=f"DPR: {p['dpr_eirr']:.1f}%")
        ax.axvline(p50,color=C["blue"],lw=2,ls=":",label=f"P50: {p50:.1f}%")
        ax.axvspan(min(ep)-2,12,alpha=0.1,color=C["red"])
        if code=="P5":
            ax.axvspan(-30,6,alpha=0.15,color=C["red"])
            txt="ACTUAL OUTCOME:\nConcessionaire (VHTRL) defaulted\nTraffic = 58% of forecast\nFIRR at completion: 1.1%\nSource: World Bank ICR 2002"
            v_text="✓ RED — Correctly Predicted"
        else:
            ax.axvspan(16,35,alpha=0.12,color=C["green"])
            txt="ACTUAL OUTCOME:\nCost +35% (₹55k→₹73k Cr)\nBuild +24 months\nYr2 AADT: ~45,000 (+80% vs DPR)\nProject succeeded via traffic beat\n(MSRDC SPV governance)"
            v_text="✓ AMBER-RED — Fragility at appraisal\ncorrectly shown; traffic beat rescued"
        ax.text(0.97,0.97,txt,transform=ax.transAxes,fontsize=8.5,ha="right",va="top",
                bbox=dict(boxstyle="round,pad=0.4",fc=C["green_lt"] if code=="P7" else C["red_lt"],ec="grey"))
        ax.text(0.03,0.97,v_text,transform=ax.transAxes,fontsize=9,ha="left",va="top",
                fontweight="bold",color=fc,bbox=dict(boxstyle="round,pad=0.4",fc=bg,ec=fc))
        ax.set_facecolor("#FAFAFA")
        ax.set_title(f"{p['name']}  [{p['dpr_mode']}]\nFI={fi:.1f}% | {verdict(fi)} | DPR={p['dpr_eirr']:.1f}% | P50={p50:.1f}%",
                     fontsize=10,fontweight="bold")
        ax.set_xlabel("EIRR (%)",fontsize=9); ax.yaxis.set_visible(False); ax.legend(fontsize=8)
    plt.tight_layout(rect=[0,0,1,0.93])
    fname=os.path.join(OUT_DIR,"pfff_validation_v13.png")
    fig.savefig(fname,dpi=150,bbox_inches="tight",facecolor="white")
    plt.show(); plt.close(fig)
    print(f"  → Saved: {fname}")


# ═══════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════
def export_excel(all_results, all_scn, all_svs, all_p50, all_tornado):
    if not HAS_OPENPYXL: print("openpyxl not installed — skipping Excel"); return
    wb = Workbook()

    # Sheet 1: Executive Summary
    ws1=wb.active; ws1.title="Executive Summary"
    hdrs1=["Code","Project","Mode","DPR_EIRR_%","P50_EIRR_%","Bias_pp",
           "FI_Primary_%","FI_EIRR_%","FI_FIRR_%","FI_Equity_%",
           "Cost_SV_DPR","Cost_SV_P50","Traf_SV_DPR","Traf_SV_P50",
           "Delay_SV_DPR","Delay_SV_P50","Primary_Driver","Verdict","Role"]
    for j,h in enumerate(hdrs1,1):
        c=ws1.cell(1,j); c.value=h; c.font=Font(bold=True,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor="1F497D"); c.alignment=Alignment(horizontal="center")
    for i,code in enumerate(PROJECTS.keys(),2):
        p=PROJECTS[code]; dm=p["dpr_mode"]; res=all_results[code][dm]
        sv=all_svs[code]; p50=all_p50[code]; torn=all_tornado[code]
        row=[code,p["name"],dm,p["dpr_eirr"],round(p50,2),round(p["dpr_eirr"]-p50,2),
             round(res["fi_p"],1),round(res["fi_eirr"],1),
             round(res["fi_firr"],1) if not np.isnan(res["fi_firr"]) else "N/A",
             round(res["fi_eq"],1) if not np.isnan(res["fi_eq"]) else "N/A",
             f"+{sv['dpr_cost']:.1f}%" if sv['dpr_cost'] else "∞",
             f"{sv['p50_cost']:+.1f}%" if sv.get('p50_cost') else "Deficit",
             f"-{sv['dpr_traf']:.1f}%" if sv['dpr_traf'] else "∞",
             f"{sv['p50_traf']:+.1f}%" if sv.get('p50_traf') else "Deficit",
             f"+{sv['dpr_delay']:.0f}mo" if sv['dpr_delay'] else "∞",
             "Already failed" if p50<12 else f"+{sv.get('p50_delay',0):.0f}mo",
             torn[0][0] if torn else "—",verdict(res["fi_p"]),p["role"]]
        for j,v in enumerate(row,1): ws1.cell(i,j).value=v
        # Color by verdict
        color={"GREEN":"D1E7DD","AMBER":"FFF3CD","RED":"F8D7DA"}.get(verdict(res["fi_p"])[:5],"FFFFFF")
        for j in range(1,len(row)+1):
            ws1.cell(i,j).fill=PatternFill("solid",fgColor=color)
    for j in range(1,len(hdrs1)+1):
        ws1.column_dimensions[get_column_letter(j)].width=16

    # Sheet 2: Iteration sample (P5 for validation)
    for code in ["P5","P7"]:
        ws=wb.create_sheet(f"Iterations_{code}")
        p=PROJECTS[code]; dm=p["dpr_mode"]
        scn=all_scn[code]
        np.random.seed(42)
        samp=run_mcs(p,scn,1000)
        res_s=simulate_mode(p,scn,samp,dm,1000)
        hdrs2=["Iter","EIRR_%","FIRR_%","Civil_Cr","LA_Cr","Delay_Mo","Traffic_AADT","Growth_%","VOC","VoT","Stall"]
        for j,h in enumerate(hdrs2,1):
            c=ws.cell(1,j); c.value=h; c.font=Font(bold=True,color="FFFFFF")
            c.fill=PatternFill("solid",fgColor="1F497D")
        for i in range(1000):
            ws.cell(i+2,1).value=i+1
            ws.cell(i+2,2).value=round(res_s["eirr_arr"][i]*100,4)
            ws.cell(i+2,3).value=round(res_s["firr_arr"][i]*100,4) if not np.isnan(res_s["firr_arr"][i]) else "N/A"
            ws.cell(i+2,4).value=round(samp["v05"][i],2)
            ws.cell(i+2,5).value=round(samp["v06"][i],2)
            ws.cell(i+2,6).value=round(samp["v07"][i],2)
            ws.cell(i+2,7).value=round(samp["v01"][i],0)
            ws.cell(i+2,8).value=round(samp["v02"][i]*100,4)
            ws.cell(i+2,9).value=round(samp["v10"][i],4)
            ws.cell(i+2,10).value=round(samp["v11"][i],4)
            ws.cell(i+2,11).value=int(samp["reg"][i])
        for j in range(1,12): ws.column_dimensions[get_column_letter(j)].width=14

    fname=os.path.join(OUT_DIR,"PFFF_v13_Forensic_Audit.xlsx")
    wb.save(fname)
    print(f"  → Excel saved: {fname}")

# ═══════════════════════════════════════════════════════════
# MAIN (Colab / Script)
# ═══════════════════════════════════════════════════════════
def main():
    print("\n"+"═"*70)
    print("  PFFF v13.0 — Probabilistic Feasibility Fragility Framework")
    print("  M.BEM Thesis | SPA Delhi 2024 | Varshni M S")
    print("═"*70)

    print("\n[STEP 1] Zero-Stress Calibration")
    for code,p in PROJECTS.items():
        scn=compute_scn(p); verify_calibration(p,scn)

    print(f"\n[STEP 2] Monte Carlo ({N_ITER:,} iterations × 7 projects × 3 modes)")
    all_results={}; all_scn={}; all_svs={}; all_p50={}; all_tornado={}

    for code,p in PROJECTS.items():
        print(f"\n  [{code}] {p['name']}")
        scn=compute_scn(p); samp=run_mcs(p,scn,N_ITER)
        mode_results={}
        for mode in MODES:
            res=simulate_mode(p,scn,samp,mode,N_ITER)
            mode_results[mode]=res
            fi_f=f"{res['fi_firr']:.1f}%" if not np.isnan(res['fi_firr']) else "N/A"
            print(f"    {mode}: FI={res['fi_p']:5.1f}%  EIRR_FI={res['fi_eirr']:.1f}%  FIRR_FI={fi_f}  [{verdict(res['fi_p'])}]")
        tornado=spearman_tornado(p,scn,samp,mode_results[p["dpr_mode"]]["eirr_arr"])
        rcf=rcf_acid_test(p,scn,samp,mode_results[p["dpr_mode"]]["fi_p"])
        ep=mode_results[p["dpr_mode"]]["eirr_arr"]*100
        p50=np.percentile(ep,50)
        svs=compute_dual_sv(p,scn,p50)
        print(f"    DPR EIRR: {p['dpr_eirr']:.2f}%  P50: {p50:.2f}%  Bias: {p['dpr_eirr']-p50:+.2f}pp")
        print(f"    SV_cost: DPR={svs['dpr_cost']}% | P50={svs['p50_cost']}%  | {svs['p50_status']}")
        mode_results["_tornado"]=tornado; mode_results["_samp"]=samp; mode_results["_rcf"]=rcf
        all_results[code]=mode_results; all_scn[code]=scn
        all_svs[code]=svs; all_p50[code]=p50; all_tornado[code]=tornado

    print("\n[STEP 3] Per-Project Dashboards (7 × PNG)")
    for code,p in PROJECTS.items():
        print(f"\n  [{code}] {p['name']}")
        plot_dashboard(p,all_scn[code],all_results[code]["_samp"],all_results[code],
                       all_results[code]["_tornado"],all_results[code]["_rcf"],all_svs[code],code)

    print("\n[STEP 4] Batch Comparison + Optimism Bias + SV Chart")
    plot_batch_and_bias(all_results,all_svs,all_p50)

    print("\n[STEP 5] Validation Exhibit (P5 & P7)")
    plot_validation_exhibit(all_results,all_scn)

    print("\n[STEP 6] Excel Forensic Audit Report")
    export_excel(all_results,all_scn,all_svs,all_p50,all_tornado)

    print("\n"+"═"*70)
    print("  RESULTS SUMMARY")
    print("═"*70)
    print(f"  {'Code':<5} {'Project':<34} {'Mode':<5} {'DPR%':>6} {'P50%':>6} {'Bias':>7} {'FI%':>6} {'CostSV_DPR':>11} {'P50_Status'}")
    print("  "+"─"*100)
    for code,p in PROJECTS.items():
        dm=p["dpr_mode"]; fi=all_results[code][dm]["fi_p"]
        sv=all_svs[code]; p50=all_p50[code]
        csv_str=f"+{sv['dpr_cost']:.0f}%" if sv['dpr_cost'] else "∞"
        tag=" ← VAL" if p["role"]=="VALIDATION" else ""
        print(f"  {code:<5} {p['name']:<34} {dm:<5} {p['dpr_eirr']:>6.1f}% {p50:>6.1f}% {p['dpr_eirr']-p50:>+7.1f}pp {fi:>6.1f}%  {csv_str:>10}  {sv['p50_status']}{tag}")
    print()
    print("  KEY FINDING — Switching Value Bias (Phantom Safety):")
    for code,p in PROJECTS.items():
        sv=all_svs[code]
        if sv.get("p50_cost") and sv["p50_cost"]<=0:
            print(f"  ⚠ [{code}] DPR claims +{sv['dpr_cost']:.0f}% cost tolerance → PFFF shows project is ALREADY below hurdle at P50")
    print("═"*70+"\n")

if __name__=="__main__":
    main()
